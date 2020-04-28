# Modyfikuje cene w arkuszu excel, zmniejszajac ja o 10% i dodajemy wykres

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)  # otwieramy plik excel
    sheet = wb['Sheet1']   # Wybieramy, ktory arkusz
    cell = sheet['a1']  # Wybieramy, ktora komorka
    # cell = sheet.cell(1,1) Gdzie (1,1) to kolumna i wiersz. To tutaj jest tym samym co wyzej
    # print(cell.value)  # Pokazuje, jaka wartosc jest w komorce
    # Teraz musimy tak przeleciec przez wszystkie wiersze, podac komorke, w ktorej jest cena, a nastepnie chcemy
    # zmniejszyc cene o 10%. Piersze co musimy wiedziec to ile mamy wierszy w "transactions_id"

    # print(sheet.max_row)  Ta komenda mozemy sprawdzic ile ma wierszy

    # Zatem musimy zrobic tak, aby przeleciec przez wszystkie wiersze. Loopem ofc
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)   # sheet.cell aby dostac dostep do komorki "price"
        corrected_price = cell.value * 0.9
    # Chce aby wszystkie nowe ceny znalazly sie w kolumnie "D"
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # Zapisuje, ale w osobnym pliku 'in case' program ma buga ^^
    wb.save('transactions2.xlsx')     # Tutaj komenda, jak zapisac

    # Do tego momentu wszystko ladnie smiga. Teraz dodaje wykres, a wiec importowanko
    values = Reference(sheet,       # W zminnej values bedziemy mieli wszystkie te wartosci z komorki "D"
              min_row=2,
              max_row =sheet.max_row,
              min_col =4,
              max_col =4)

    chart = BarChart()  # Tworzymy wykres
    chart.add_data(values)  # Co ma dodac
    sheet.add_chart(chart, 'a7')  # Miejsce, w ktore dodajemy wykres

    wb.save(filename)

    # Jesli chcielibysmy uzyc tego programu do automatycznego update'owania pliku to po prostu by nie dzialal
    # Zeby dzialal musimy go przerobic na funkcje, ktora bedzie sama to robila

